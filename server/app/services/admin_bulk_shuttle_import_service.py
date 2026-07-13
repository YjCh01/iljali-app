"""어드민 — 엑셀 셔틀 노선 일괄 등록."""

from __future__ import annotations

import io
import re
from uuid import uuid4

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.services.admin_ops_service import _audit
from app.services.entitlement_service import normalize_brn
from app.services.shuttle_commute_service import list_routes_for_company, upsert_commute_route

_KAKAO_ADDRESS_URL = "https://dapi.kakao.com/v2/local/search/address.json"
_WORKPLACE_STOP_ID = "__shuttle_workplace__"
_WORKPLACE_LABEL = "근무지"
_DEFAULT_LAT = 37.5012
_DEFAULT_LNG = 127.0396
_MAX_ROUTES = 30
_MAX_STOPS_PER_ROUTE = 15
_OVERLAY_COLORS = ("#E53935", "#1E88E5", "#43A047", "#FB8C00", "#8E24AA")

_HEADER_ALIASES = {
    "route_name": ("노선명", "route_name", "route name", "노선"),
    "order": ("정류장순서", "순서", "order", "stop_order"),
    "stop_name": ("정류장명", "정류장", "stop_name", "stop name"),
    "time": ("도착시간", "시간", "time", "도착/출발시간", "출발시간"),
    "address": ("주소", "address", "도로명주소"),
}

_DATA_SHEET_NAMES = ("노선입력", "routes", "data")
_SKIP_SHEET_NAMES = ("작성안내", "guide", "instructions")


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _resolve_columns(header_row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        key = _normalize_header(cell)
        if not key:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if key in {_normalize_header(alias) for alias in aliases}:
                mapping[field] = index
                break
    required = ("route_name", "order", "stop_name", "time")
    missing = [field for field in required if field not in mapping]
    if missing:
        raise ValueError(
            "엑셀 헤더가 올바르지 않습니다. "
            "노선명·정류장순서·정류장명·도착시간 열이 필요합니다."
        )
    return mapping


def _cell(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _normalize_time(raw: str) -> str | None:
    text = (raw or "").strip()
    if not text:
        return None
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        parts = text.split(":")
        hour = int(parts[0])
        minute = int(parts[1])
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
        return None
    if re.fullmatch(r"\d{3,4}", text):
        padded = text.zfill(4)
        hour = int(padded[:2])
        minute = int(padded[2:])
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
    return None


def _geocode(query: str) -> tuple[float, float] | None:
    keyword = (query or "").strip()
    if not keyword or not settings.kakao_rest_api_key:
        return None
    try:
        with httpx.Client(timeout=10.0) as client:
            response = client.get(
                _KAKAO_ADDRESS_URL,
                params={"query": keyword},
                headers={"Authorization": f"KakaoAK {settings.kakao_rest_api_key}"},
            )
            if response.status_code >= 400:
                return None
            documents = response.json().get("documents") or []
            if not documents:
                return None
            doc = documents[0]
            return float(doc["y"]), float(doc["x"])
    except (httpx.HTTPError, KeyError, TypeError, ValueError):
        return None


def _pick_data_sheet(workbook) -> object:
    for name in _DATA_SHEET_NAMES:
        if name in workbook.sheetnames:
            return workbook[name]
    for name in workbook.sheetnames:
        if name.strip().lower() in _SKIP_SHEET_NAMES:
            continue
        return workbook[name]
    raise ValueError("엑셀에 입력 시트가 없습니다.")


def _parse_rows(file_bytes: bytes) -> list[dict]:
    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = _pick_data_sheet(workbook)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("엑셀에 데이터가 없습니다.")

    header_index = None
    columns: dict[str, int] | None = None
    for index, row in enumerate(rows):
        if not any(cell not in (None, "") for cell in row):
            continue
        try:
            columns = _resolve_columns(row)
            header_index = index
            break
        except ValueError:
            continue
    if columns is None or header_index is None:
        raise ValueError(
            "엑셀 헤더를 찾을 수 없습니다. "
            "첫 줄에 노선명·정류장순서·정류장명·도착시간을 넣어 주세요."
        )

    parsed: list[dict] = []
    for row in rows[header_index + 1 :]:
        route_name = _cell(row, columns["route_name"])
        stop_name = _cell(row, columns["stop_name"])
        if not route_name and not stop_name:
            continue
        if not route_name or not stop_name:
            raise ValueError("노선명과 정류장명은 모두 입력해야 합니다.")
        order_raw = _cell(row, columns["order"])
        if not order_raw.isdigit():
            raise ValueError(f'"{route_name}" · "{stop_name}" — 정류장순서는 숫자여야 합니다.')
        time_raw = _cell(row, columns["time"])
        normalized_time = _normalize_time(time_raw)
        if normalized_time is None:
            raise ValueError(
                f'"{route_name}" · "{stop_name}" — '
                f'도착시간은 HH:MM 형식이어야 합니다. (입력: {time_raw or "비어 있음"})'
            )
        parsed.append(
            {
                "route_name": route_name,
                "order": int(order_raw),
                "stop_name": stop_name,
                "time": normalized_time,
                "address": _cell(row, columns.get("address")),
            }
        )
    if not parsed:
        raise ValueError("등록할 정류장 행이 없습니다.")
    return parsed


def _group_routes(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        grouped.setdefault(row["route_name"], []).append(row)
    if len(grouped) > _MAX_ROUTES:
        raise ValueError(f"한 번에 등록 가능한 노선은 최대 {_MAX_ROUTES}개입니다.")
    for route_name, stops in grouped.items():
        if len(stops) > _MAX_STOPS_PER_ROUTE:
            raise ValueError(f'"{route_name}" — 정류장은 노선당 최대 {_MAX_STOPS_PER_ROUTE}개입니다.')
        if len(stops) < 2:
            raise ValueError(
                f'"{route_name}" — 정류장은 최소 2곳(경유 1 + 근무지 1)이 필요합니다.'
            )
        orders = [stop["order"] for stop in stops]
        if len(set(orders)) != len(orders):
            raise ValueError(f'"{route_name}" — 정류장순서가 중복되었습니다.')
        stops.sort(key=lambda item: item["order"])
    return grouped


def _build_route_payload(
    *,
    company_key: str,
    route_name: str,
    stops: list[dict],
    route_id: str,
    overlay_color_hex: str,
) -> tuple[dict, list[str]]:
    warnings: list[str] = []
    built_stops: list[dict] = []
    last_index = len(stops) - 1

    for index, stop in enumerate(stops):
        geocode_query = stop["address"] or stop["stop_name"]
        coord = _geocode(geocode_query)
        if coord is None:
            coord = (_DEFAULT_LAT, _DEFAULT_LNG)
            warnings.append(
                f'"{route_name}" · "{stop["stop_name"]}" — '
                f"주소 좌표를 찾지 못해 기본 위치를 사용했습니다."
            )
        latitude, longitude = coord
        is_workplace = index == last_index
        if is_workplace:
            built_stops.append(
                {
                    "id": _WORKPLACE_STOP_ID,
                    "label": _WORKPLACE_LABEL,
                    "coordinate": {"latitude": latitude, "longitude": longitude},
                    "arrivalTime": stop["time"],
                }
            )
        else:
            built_stops.append(
                {
                    "id": f"stop_{stop['order']}_{uuid4().hex[:8]}",
                    "label": stop["stop_name"],
                    "coordinate": {"latitude": latitude, "longitude": longitude},
                    "departureTime": stop["time"],
                }
            )

    route = {
        "id": route_id,
        "companyKey": company_key,
        "routeName": route_name,
        "active": True,
        "overlayColorHex": overlay_color_hex,
        "stops": built_stops,
        "polylinePoints": [],
    }
    return route, warnings


def bulk_import_shuttle_routes_from_excel(
    db: Session,
    *,
    company_key: str,
    file_bytes: bytes,
    replace_existing: bool = True,
) -> dict:
    brn = normalize_brn(company_key)
    if not brn:
        raise ValueError("company_key(사업자번호)가 필요합니다.")

    rows = _parse_rows(file_bytes)
    grouped = _group_routes(rows)
    existing = list_routes_for_company(db, company_key=brn, active_only=False)
    existing_by_name = {
        str(route.get("routeName", "")).strip(): route
        for route in existing
        if str(route.get("routeName", "")).strip()
    }

    results: list[dict] = []
    imported = 0
    updated = 0

    for color_index, (route_name, stops) in enumerate(grouped.items()):
        existing_route = existing_by_name.get(route_name)
        if existing_route and replace_existing:
            route_id = str(existing_route.get("id", "")).strip() or f"route_{uuid4().hex[:12]}"
            action = "updated"
        else:
            route_id = f"route_{uuid4().hex[:12]}"
            action = "created"

        route_payload, warnings = _build_route_payload(
            company_key=brn,
            route_name=route_name,
            stops=stops,
            route_id=route_id,
            overlay_color_hex=_OVERLAY_COLORS[color_index % len(_OVERLAY_COLORS)],
        )
        saved = upsert_commute_route(db, route=route_payload)
        if action == "updated":
            updated += 1
        else:
            imported += 1
        results.append(
            {
                "route_name": route_name,
                "route_id": saved.get("id", route_id),
                "action": action,
                "stop_count": len(stops),
                "warnings": warnings,
            }
        )

    _audit(
        db,
        action="shuttle.bulk_import",
        target_type="company",
        target_id=brn,
        detail={
            "submitted_routes": len(grouped),
            "imported": imported,
            "updated": updated,
        },
    )
    db.commit()

    return {
        "company_key": brn,
        "submitted_routes": len(grouped),
        "imported": imported,
        "updated": updated,
        "results": results,
    }
