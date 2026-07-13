"""셔틀 노선 엑셀 입력 양식 생성."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "assets" / "templates" / "shuttle_route_import_template.xlsx"

HEADERS = ("노선명", "정류장순서", "정류장명", "도착시간", "주소")
SAMPLE_ROWS = [
    ("1호차", 1, "강남역 2번출구", "06:30", "서울특별시 강남구 강남대로 396"),
    ("1호차", 2, "삼성역 5번출구", "06:45", "서울특별시 강남구 테헤란로 521"),
    ("1호차", 3, "근무지", "07:30", "경기도 이천시 마장면"),
    ("2호차", 1, "수원역 버스환승센터", "05:50", "경기도 수원시 팔달구"),
    ("2호차", 2, "근무지", "06:40", "경기도 이천시 마장면"),
]

GUIDE_LINES = [
    "【셔틀 노선 엑셀 작성 안내】",
    "",
    "1. '노선입력' 시트에 아래 5개 열을 채워 주세요.",
    "   · 노선명 — 같은 이름끼리 한 노선으로 묶입니다 (예: 1호차, A노선)",
    "   · 정류장순서 — 1부터 순서대로, 마지막 순서 = 근무지(도착지)",
    "   · 정류장명 — 정류장 이름 (마지막은 '근무지' 권장)",
    "   · 도착시간 — HH:MM (24시간). 경유 정류장=탑승 시각, 마지막=근무지 도착 시각",
    "   · 주소 — 도로명주소 권장 (지도 핀 위치). 없으면 정류장명으로 좌표 검색",
    "",
    "2. 노선당 정류장 2곳 이상 (경유 1 + 근무지 1), 최대 15곳",
    "",
    "3. 작성 후 어드민 → 회원 관리 → 기업 선택 → 엑셀 업로드",
    "",
    "※ 예시는 '노선입력' 시트 2~6행 — 삭제 후 실제 데이터를 입력하세요.",
]


def _style_header_row(sheet) -> None:
    fill = PatternFill("solid", fgColor="E8EAF6")
    font = Font(bold=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autosize(sheet) -> None:
    widths = (14, 12, 24, 12, 36)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook() -> Workbook:
    workbook = Workbook()
    data = workbook.active
    data.title = "노선입력"
    _style_header_row(data)
    for row_index, row in enumerate(SAMPLE_ROWS, start=2):
        for col_index, value in enumerate(row, start=1):
            data.cell(row=row_index, column=col_index, value=value)
    _autosize(data)

    guide = workbook.create_sheet("작성안내")
    for row_index, line in enumerate(GUIDE_LINES, start=1):
        cell = guide.cell(row=row_index, column=1, value=line)
        if row_index == 1:
            cell.font = Font(bold=True, size=13)
    guide.column_dimensions["A"].width = 88
    return workbook


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
